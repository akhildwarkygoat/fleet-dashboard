#!/usr/bin/env python3
"""Recreate the ERP's DATED vehicle costing as an Excel workbook.

Companion to build_cost_sheet.py. That script reads the employee/attendance feed
(/api/general/VehicleEmpMapDetails), whose cost lines carry no date — which is why
its workbook could not say which year a payment belonged to. This one reads

    POST /api/general/VehicleEmpMapProjectDetails   (body {})

the vehicle PROJECT feed, where every cost line is a dated, approved plan entry:

  Proj_Activity_Name  the cost head        (ROAD TAX, VEHICLE INSURANCE, TYRE, ...)
  Description         the finer detail     (MAINTENANCE SERVICE - 2, GREEN TAX, ...)
  Period_Name         YEARLY / QUARTER - I..IV / a month name
  From_Date/To_Date   the period the amount actually covers  <- the dates the other feed lacks
  Rate                the UNIT rate (per litre, per tyre, per job)
  Pur_Amount          rate x quantity = the amount actually purchased; 0 until approved

Usage:  python build_project_cost_sheet.py [project.json] [out.xlsx] [live_erp.json]
"""
import json, sys, collections
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "data/erp_project_live.json"
OUT = sys.argv[2] if len(sys.argv) > 2 else "ERP vehicle project costing.xlsx"
LIVE = sys.argv[3] if len(sys.argv) > 3 else "data/erp_live.json"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
NOTE_FONT = Font(name=FONT, italic=True, size=9, color="7F7F7F")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
INPUT_FONT = Font(name=FONT, size=10, bold=True, color="0000FF")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
MONEY = '#,##0.00;(#,##0.00);-'
DATEF = 'dd-mm-yyyy'
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def pdate(s):
    """'05-07-2024 00:00:00' -> date(2024, 7, 5). Blank -> None."""
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%d-%m-%Y").date()
    except ValueError:
        return None


def fy_of(d):
    """Indian financial year label for a date: 05-07-2024 -> '2024-25'."""
    if not d:
        return ""
    y = d.year if d.month >= 4 else d.year - 1
    return f"{y}-{(y + 1) % 100:02d}"


def num(s):
    try:
        return float(s or 0)
    except (TypeError, ValueError):
        return 0.0


def header(ws, cols, row=1):
    for i, (title, width) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=title)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def main():
    rows = json.load(open(SRC, encoding="utf-8"))
    for r in rows:
        r["_from"], r["_to"] = pdate(r["From_Date"]), pdate(r["To_Date"])
        r["_ord"], r["_appr"] = pdate(r["Order_Date"]), pdate(r["Approved_Date"])
        r["_fy"] = fy_of(r["_from"])
    rows.sort(key=lambda r: (r["Veh_Name"], r["_from"] or date(1900, 1, 1), r["Proj_Activity_Name"]))

    vehicles = sorted({r["Veh_Name"] for r in rows})
    acts = sorted({r["Proj_Activity_Name"] for r in rows})
    fys = sorted({r["_fy"] for r in rows if r["_fy"]})
    LATEST_FY = fys[-1]
    # default as-of is today, so "Active" reads as active now; outside the feed's own span
    # (an old extract, or one built before the year starts) fall back to its newest period start
    span_lo = min(r["_from"] for r in rows if r["_from"])
    span_hi = max(r["_to"] for r in rows if r["_to"])
    ASOF = date.today() if span_lo <= date.today() <= span_hi else max(r["_from"] for r in rows if r["_from"])
    CY = ASOF.year
    # a vehicle can appear under more than one division; take its commonest
    div_of = {v: collections.Counter(r["Division"] for r in rows if r["Veh_Name"] == v).most_common(1)[0][0]
              for v in vehicles}

    try:
        live = json.load(open(LIVE, encoding="utf-8"))
        fleet = {}
        for r in live:
            fleet.setdefault(r["VehName"], {"Type": r.get("Type"), "Seat": r.get("Seat"),
                                            "Compname": r.get("Compname")})
    except (OSError, ValueError):
        fleet = {}

    wb = Workbook()
    LAST = len(rows) + 1                      # last data row on CostLines
    R = lambda col: f"CostLines!${col}$2:${col}${LAST}"   # noqa: E731

    # ---------------------------------------------------------------- README
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 104
    ws["A1"] = "ERP vehicle project costing — dated cost lines"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)

    ws["A3"] = "CONTROLS — edit the yellow cells; every sheet follows"
    ws["A3"].font = Font(name=FONT, bold=True, size=11, color="1F3864")
    for r_, label, val, fmt in ((4, "As-of date", ASOF, DATEF),
                                (5, "Calendar year", CY, "0"),
                                (6, "Financial year", LATEST_FY, "@")):
        ws.cell(row=r_, column=1, value=label).font = BOLD
        c = ws.cell(row=r_, column=2, value=val)
        c.font, c.fill, c.number_format, c.border = INPUT_FONT, INPUT_FILL, fmt, BOX
    ws["C4"] = "drives the 'Active' flag on CostLines and the 'Active' column on ByVehicle"
    ws["C5"] = "drives the 'In calendar year' flag on CostLines"
    ws["C6"] = f"drives every FY column on ByVehicle. Financial years present: {', '.join(fys)}"
    for r_ in (4, 5, 6):
        ws.cell(row=r_, column=3).font = NOTE_FONT

    total_pur = sum(num(r["Pur_Amount"]) for r in rows)
    unappr = [r for r in rows if not r["_appr"]]
    meta = [
        ("", ""),
        ("Source", "POST http://172.16.10.169:8089/api/general/VehicleEmpMapProjectDetails  (body {}, Content-Type: application/json)"),
        ("Pulled", f"{len(rows):,} rows, all Project_Type = VEHICLE"),
        ("Vehicles", f"{len(vehicles)} distinct"),
        ("Cost heads", f"{len(acts)} Proj_Activity_Name values / {len({r['Description'] for r in rows})} Description values"),
        ("Period covered", f"{min(r['_from'] for r in rows if r['_from']).strftime('%d-%m-%Y')} to "
                           f"{max(r['_to'] for r in rows if r['_to']).strftime('%d-%m-%Y')}  ({', '.join(fys)})"),
        ("Total Pur_Amount", f"Rs {total_pur:,.0f} across all years"),
        ("", ""),
        ("WHY THIS FEED", ""),
        ("It is dated", "The employee feed's cost lines carry no date, so no payment could be assigned to a year — the open "
                        "question at the end of 'ERP vehicle costing.xlsx'. Here every line has From_Date/To_Date and a "
                        "Period_Name, so a cost can be placed in a financial year and pro-rated to a day."),
        ("", ""),
        ("HOW THE ERP MODELS IT", ""),
        ("Proj_Activity_Name", "The cost head: " + ", ".join(acts)),
        ("Description", "A finer split of the same line — VEHICLE OUTSIDE SERVICES breaks into MAINTENANCE SERVICE, "
                        "MAINTENANCE SERVICE - 2/3, GLASS STICKERS and so on. Equal to Proj_Activity_Name on "
                        f"{sum(1 for r in rows if r['Description'] == r['Proj_Activity_Name']):,} of {len(rows):,} rows."),
        ("Period_Name", "YEARLY, QUARTER - I..IV, or a month name. The amount covers THAT period only, so lines are not "
                        "comparable until they are put on the same footing — use the Days column to pro-rate."),
        ("Rate is a UNIT rate", f"Rate is the rate per unit and Pur_Amount is rate x quantity — they are equal only where the "
                                f"quantity is 1. Pur_Amount is the amount to sum; Rate never is. On the "
                                f"{sum(1 for r in rows if abs(num(r['Rate']) - num(r['Pur_Amount'])) > 0.01 and num(r['Pur_Amount'])) } "
                                "rows where they differ, ADBLU is Rs 60-61 per litre and TYRE is a per-tyre rate. Summing Rate "
                                "instead of Pur_Amount would understate AdBlue alone by about 40x."),
        ("Quantity is derived", "The feed has no quantity field, so CostLines derives it as Pur_Amount / Rate. It comes out a "
                                "whole number on 1,028 of the 1,029 priced rows — 45 litres, 6 tyres — which is what confirms "
                                "the reading. The exception is TN57AC3636 RTO EXPENSE at 1.87x (Rs 7,500 rate, Rs 14,000 "
                                "purchased); that one row is either a part-charge or an error, and is worth asking about."),
        ("Approved_Date", f"Blank on {len(unappr)} rows, and those are exactly the rows where Pur_Amount is 0 "
                          f"(Rs {sum(num(r['Rate']) for r in unappr):,.0f} of rates planned, none of it purchased). Approval is "
                          "what turns a planned line into spend. Every FY2026-27 line is approved."),
        ("Grs_Amount", "Identical to Pur_Amount on all rows in this pull, so it is not carried into CostLines."),
        ("Employee / Testing", "Constant across the feed (VIVEK B / 'T') — the record's owner and a status flag, not "
                               "per-vehicle data. Kept on CostLines verbatim but carries no information here."),
        ("", ""),
        ("WHAT IT DOES NOT CARRY", ""),
        ("No diesel, no salary", "The cost heads are tax, insurance, tyres, servicing, RTO and FC work only. Fuel and driver "
                                 "salary — the two largest running costs — are in neither this feed nor the employee feed, "
                                 "so a per-bus running cost cannot be built from the ERP alone yet."),
        ("Rented buses have none", f"{sum(1 for v in fleet if v not in set(vehicles))} of the {len(fleet)} vehicles in the live "
                                   "fleet have no line here, and they are exactly the RENT VEHICLE entries. Every "
                                   "COMPANY VEHICLE is covered. See the FleetCoverage sheet — the hire charge for a rented "
                                   "bus is invoiced, not planned, so it lives somewhere else in the ERP."),
        ("ADBLU stops in 2024", f"{sum(1 for r in rows if r['Proj_Activity_Name'] == 'ADBLU')} AdBlue lines exist and every one "
                                "starts in FY2024-25. Either the head was retired or later AdBlue is being booked elsewhere; "
                                "worth asking before treating a 2026 total as complete."),
        ("", ""),
        ("READ WITH CARE", ""),
        ("Repeated lines are real", f"{len(rows) - len({(r['Veh_Name'], r['Proj_Activity_Name'], r['Period_Name'], r['From_Date'], r['Pur_Amount']) for r in rows}):,} "
                                    "rows repeat an existing (vehicle, head, period, from-date, amount) combination. Unlike the "
                                    "employee feed these are not a join artefact — a bus can genuinely take two tyre purchases "
                                    "in one period — but a duplicated entry would look identical, so large repeats deserve a check."),
        ("Not a full-year view", "A period that has not started yet has no row. FY totals for the current year grow as the year "
                                 "is planned out, so they are not comparable with a closed year until it closes."),
    ]
    r_ = 8
    for k, v in meta:
        if k and not v:
            ws.cell(row=r_, column=1, value=k).font = Font(name=FONT, bold=True, size=11, color="1F3864")
        elif k:
            ws.cell(row=r_, column=1, value=k).font = BOLD
            c = ws.cell(row=r_, column=2, value=v)
            c.font = BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
        r_ += 1

    # ------------------------------------------------------------- CostLines
    ws = wb.create_sheet("CostLines")
    header(ws, [("Veh_Name", 14), ("Division", 11), ("Proj_Activity_Name", 26), ("Description", 26),
                ("Period_Name", 14), ("From_Date", 12), ("To_Date", 12), ("Days", 7), ("FY", 10),
                ("Rate (per unit)", 14), ("Qty (derived)", 12), ("Pur_Amount (actual)", 16), ("Per day", 11),
                ("Approved_Date", 14), ("Approved", 10), ("Active on as-of", 14),
                ("In cal. year", 12), ("Order_No", 14), ("Order_Date", 12), ("Plan_Entry_No", 13),
                ("Project_Type", 13), ("Testing", 9), ("Employee", 14)])
    for i, rec in enumerate(rows, start=2):
        vals = [rec["Veh_Name"], rec["Division"], rec["Proj_Activity_Name"], rec["Description"],
                rec["Period_Name"], rec["_from"], rec["_to"],
                f"=IF(OR($F{i}=\"\",$G{i}=\"\"),\"\",$G{i}-$F{i}+1)",
                f"=IF($F{i}=\"\",\"\",IF(MONTH($F{i})>=4,YEAR($F{i})&\"-\"&TEXT(YEAR($F{i})+1-2000,\"00\"),"
                f"YEAR($F{i})-1&\"-\"&TEXT(YEAR($F{i})-2000,\"00\")))",
                num(rec["Rate"]),
                f"=IFERROR($L{i}/$J{i},\"\")",
                num(rec["Pur_Amount"]),
                f"=IFERROR($L{i}/$H{i},\"\")",
                rec["_appr"],
                f"=IF($N{i}=\"\",\"No\",\"Yes\")",
                f"=IF(OR($F{i}=\"\",$G{i}=\"\"),0,IF(AND($F{i}<=README!$B$4,README!$B$4<=$G{i}),1,0))",
                f"=IF(OR($F{i}=\"\",$G{i}=\"\"),0,IF(AND($F{i}<=DATE(README!$B$5,12,31),"
                f"$G{i}>=DATE(README!$B$5,1,1)),1,0))",
                rec["Order_No"], rec["_ord"], rec["Plan_Entry_No"], rec["Project_Type"],
                rec["Testing"], rec["Employee"]]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font, c.border = BODY, BOX
            if col in (6, 7, 14, 19):
                c.number_format = DATEF
            elif col in (10, 12, 13):
                c.number_format = MONEY
            elif col == 11:
                c.number_format = "#,##0.##"
            elif col in (8, 16, 17):
                c.number_format = "0"
    tr = LAST + 2
    ws.cell(row=tr, column=9, value="Total").font = BOLD
    c = ws.cell(row=tr, column=12, value=f"=SUM(L2:L{LAST})")
    c.font, c.number_format = BOLD, MONEY
    ws.cell(row=tr + 1, column=9, value="Lines").font = BOLD
    ws.cell(row=tr + 1, column=12, value=f"=COUNTA(A2:A{LAST})").font = BOLD
    ws.cell(row=tr + 3, column=1,
            value="One row = one dated cost line as the ERP returns it. Qty and 'Per day' are derived here, not ERP fields: "
                  "Qty is Pur_Amount/Rate (litres, tyres, jobs) and 'Per day' spreads the amount over its own period so "
                  "YEARLY, QUARTER and monthly lines can be compared. Rate is a unit rate — sum Pur_Amount, never Rate. "
                  "'Active on as-of' and 'In cal. year' follow the yellow control cells on README.").font = NOTE_FONT
    ws.auto_filter.ref = f"A1:W{LAST}"

    # ------------------------------------------------------------- ByVehicle
    ws = wb.create_sheet("ByVehicle")
    cols = [("Vehicle", 14), ("Division", 11)] + [(a, 16) for a in acts] + \
           [("FY total", 15), ("FY lines", 10), ("Active on as-of", 15),
            ("All-time total", 16), ("All-time lines", 13)]
    header(ws, cols)
    first_act, last_act = 3, 2 + len(acts)
    for i, v in enumerate(vehicles, start=2):
        ws.cell(row=i, column=1, value=v).font = BODY
        ws.cell(row=i, column=2, value=div_of[v]).font = BODY
        for j, a in enumerate(acts):
            col = first_act + j
            c = ws.cell(row=i, column=col,
                        value=f'=SUMIFS({R("L")},{R("A")},$A{i},{R("C")},{get_column_letter(col)}$1,'
                              f'{R("I")},README!$B$6)')
            c.font, c.number_format, c.border = BODY, MONEY, BOX
        fl, ll = get_column_letter(first_act), get_column_letter(last_act)
        for col, f in ((last_act + 1, f"=SUM({fl}{i}:{ll}{i})"),
                       (last_act + 3, f'=SUMIFS({R("L")},{R("A")},$A{i},{R("P")},1)'),
                       (last_act + 4, f'=SUMIFS({R("L")},{R("A")},$A{i})')):
            c = ws.cell(row=i, column=col, value=f)
            c.font, c.number_format, c.border = BODY, MONEY, BOX
        for col, f in ((last_act + 2, f'=COUNTIFS({R("A")},$A{i},{R("I")},README!$B$6)'),
                       (last_act + 5, f'=COUNTIFS({R("A")},$A{i})')):
            c = ws.cell(row=i, column=col, value=f)
            c.font, c.number_format = BODY, "#,##0"
    vlast = len(vehicles) + 1
    ws.cell(row=vlast + 2, column=2, value="Total").font = BOLD
    for col in range(first_act, last_act + 6):
        c = ws.cell(row=vlast + 2, column=col,
                    value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{vlast})")
        c.font = BOLD
        c.number_format = "#,##0" if col in (last_act + 2, last_act + 5) else MONEY
    ws.cell(row=vlast + 4, column=1,
            value="Every activity column is filtered to the financial year in README!B4:B6 — change the yellow cell and the "
                  "whole grid moves. 'All-time' columns ignore it. Division is the vehicle's commonest division in the feed; "
                  "a few buses appear under more than one.").font = NOTE_FONT

    # ------------------------------------------------------------ ByActivity
    ws = wb.create_sheet("ByActivity")
    header(ws, [("Proj_Activity_Name", 26)] + [(f, 16) for f in fys] +
               [("Total", 16), ("Share", 9), ("Lines", 9), ("Vehicles", 10)])
    n_fy = len(fys)
    for i, a in enumerate(acts, start=2):
        ws.cell(row=i, column=1, value=a).font = BODY
        for j, f in enumerate(fys):
            col = 2 + j
            c = ws.cell(row=i, column=col,
                        value=f'=SUMIFS({R("L")},{R("C")},$A{i},{R("I")},{get_column_letter(col)}$1)')
            c.font, c.number_format, c.border = BODY, MONEY, BOX
        tot_col = 2 + n_fy
        c = ws.cell(row=i, column=tot_col, value=f"=SUM(B{i}:{get_column_letter(tot_col - 1)}{i})")
        c.font, c.number_format = BOLD, MONEY
        s = ws.cell(row=i, column=tot_col + 1,
                    value=f"=IFERROR({get_column_letter(tot_col)}{i}/{get_column_letter(tot_col)}${len(acts) + 3},\"\")")
        s.font, s.number_format = BODY, "0.0%"
        ws.cell(row=i, column=tot_col + 2, value=f'=COUNTIFS({R("C")},$A{i})').font = BODY
        ws.cell(row=i, column=tot_col + 3,
                value=len({r["Veh_Name"] for r in rows if r["Proj_Activity_Name"] == a})).font = BODY
    alast = len(acts) + 1
    ws.cell(row=alast + 2, column=1, value="Total").font = BOLD
    for col in range(2, 2 + n_fy + 1):
        c = ws.cell(row=alast + 2, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{alast})")
        c.font, c.number_format = BOLD, MONEY
    c = ws.cell(row=alast + 2, column=2 + n_fy + 2, value=f"=SUM({get_column_letter(2 + n_fy + 2)}2:{get_column_letter(2 + n_fy + 2)}{alast})")
    c.font, c.number_format = BOLD, "#,##0"
    ws.cell(row=alast + 4, column=1,
            value="Financial years run April-March, taken from each line's From_Date. The Vehicles column is a distinct count "
                  "computed when this sheet was built, so it does not move with the other columns. The current financial year "
                  "is still being planned — treat it as partial.").font = NOTE_FONT

    # --------------------------------------------------------- FleetCoverage
    ws = wb.create_sheet("FleetCoverage")
    header(ws, [("Vehicle", 14), ("In live fleet", 12), ("Type (live)", 17), ("Seat", 6),
                ("Compname (live)", 26), ("Costed", 9), ("Cost lines", 11),
                ("FY total", 15), ("All-time total", 16)])
    allv = sorted(set(vehicles) | set(fleet))
    for i, v in enumerate(allv, start=2):
        info = fleet.get(v, {})
        ws.cell(row=i, column=1, value=v).font = BODY
        ws.cell(row=i, column=2, value="Yes" if v in fleet else "No").font = BODY
        ws.cell(row=i, column=3, value=info.get("Type")).font = BODY
        ws.cell(row=i, column=4, value=info.get("Seat")).font = BODY
        ws.cell(row=i, column=5, value=info.get("Compname")).font = BODY
        ws.cell(row=i, column=6, value=f'=IF(G{i}>0,"Yes","No")').font = BODY
        ws.cell(row=i, column=7, value=f'=COUNTIFS({R("A")},$A{i})').font = BODY
        for col, f in ((8, f'=SUMIFS({R("L")},{R("A")},$A{i},{R("I")},README!$B$6)'),
                       (9, f'=SUMIFS({R("L")},{R("A")},$A{i})')):
            c = ws.cell(row=i, column=col, value=f)
            c.font, c.number_format, c.border = BODY, MONEY, BOX
    flast = len(allv) + 1
    ws.cell(row=flast + 2, column=5, value="Total").font = BOLD
    for col in (7, 8, 9):
        c = ws.cell(row=flast + 2, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{flast})")
        c.font, c.number_format = BOLD, ("#,##0" if col == 7 else MONEY)
    ws.cell(row=flast + 4, column=1,
            value="Live fleet from data/erp_live.json (the employee feed). A RENT VEHICLE has no costing line here and a "
                  "COMPANY VEHICLE always has one — the split is exact. Vehicles marked 'No' in column B are costed but no "
                  "longer carry employees: retired or non-transport.").font = NOTE_FONT

    # ------------------------------------------------------------- RawSample
    ws = wb.create_sheet("RawSample")
    ws.column_dimensions["A"].width = 200
    ws["A1"] = "Verbatim rows from the feed — exactly as the ERP returns them"
    ws["A1"].font = Font(name=FONT, bold=True, size=11)
    ws["A2"] = "Field names and date format are the ERP's, unchanged. Dates arrive as dd-mm-yyyy strings with a 00:00:00 tail."
    ws["A2"].font = NOTE_FONT
    for n, rec in enumerate(rows[:12]):
        c = ws.cell(row=4 + n, column=1,
                    value=json.dumps({k: v for k, v in rec.items() if not k.startswith("_")}, ensure_ascii=False))
        c.font = Font(name="Courier New", size=8)

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  {len(rows):,} cost lines | {len(vehicles)} vehicles | {len(acts)} heads | FYs {', '.join(fys)}")
    print(f"  total Pur_Amount Rs {total_pur:,.0f} | as-of {ASOF} | latest FY {LATEST_FY}")


if __name__ == "__main__":
    main()
