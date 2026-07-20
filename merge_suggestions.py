#!/usr/bin/env python3
"""
merge_suggestions.py — READ-ONLY analysis for the supervisor. Changes NOTHING.

Reads the live ERP dump (data/erp_live.json), reconstructs each bus's stops exactly
as they exist in the ERP (every distinct home GPS = one stop, NO merging), then lists
which nearby stops COULD be merged — always merging the smaller stop(s) INTO the more
populated one. Runs TWO thresholds (200 m and 300 m) so they can be compared, and
writes stop_merge_suggestions.xlsx.

It does NOT touch the dashboard, current_routes.json, or any app data — it only reads
the ERP dump and writes a brand-new spreadsheet for review/approval.
"""
import json, math, argparse
from collections import Counter, defaultdict

THRESHOLDS = [200.0, 300.0]
SRC = "data/erp_live.json"
OUT = "stop_merge_suggestions.xlsx"
OUT_JSON = "public/merge_suggestions.json"   # consumed by the Prev-route "Merge review" tab
DEPOT = (10.207550, 77.806206)

_ap = argparse.ArgumentParser()
_ap.add_argument("--json-only", action="store_true",
                 help="write only public/merge_suggestions.json (skip the xlsx) — used by the live rebuild")
ARGS = _ap.parse_args()


def norm(s): return (s or "").strip()


def hav(a, b):
    R = 6371000.0
    la1, ln1, la2, ln2 = map(math.radians, (a[0], a[1], b[0], b[1]))
    d1, d2 = la2 - la1, ln2 - ln1
    h = math.sin(d1 / 2) ** 2 + math.cos(la1) * math.cos(la2) * math.sin(d2 / 2) ** 2
    return 2 * R * math.asin(math.sqrt(h))


def gps(r):
    la, ln = norm(r.get("Latitude")), norm(r.get("Longitude"))
    if not la or not ln or la == "0" or ln == "0":
        return None
    try:
        return (round(float(la), 6), round(float(ln), 6))
    except ValueError:
        return None


rows = json.load(open(SRC))
latest = sorted({norm(r.get("date")) for r in rows if norm(r.get("date"))})[-1]

# per bus -> {gps: {hc, loc Counter}} : distinct-GPS stops exactly as the ERP has them
buses = defaultdict(lambda: defaultdict(lambda: {"hc": 0, "loc": Counter()}))
unit_of, type_of = {}, {}
for r in rows:
    if norm(r.get("date")) != latest:
        continue
    v = norm(r.get("VehName") or r.get("Veh_Mas"))
    if not v:
        continue
    g = gps(r)
    if not g:
        continue
    st = buses[v][g]
    st["hc"] += 1
    loc = norm(r.get("Locality")) or norm(r.get("Village")) or norm(r.get("Area"))
    if loc:
        st["loc"][loc] += 1
    if r.get("Type"):
        type_of[v] = "rental" if "rent" in r["Type"].lower() else "owned"
    unit_of[v] = "Technotek" if "technotek" in (r.get("Compname") or "").lower() else "Gainup"


def stop_name(st):
    c = st["loc"].most_common(1)
    return c[0][0] if c else "Stop"


# stops per bus, most-populated first (computed once, reused for every threshold)
bus_stops = {}
for v in sorted(buses):
    s = [{"gps": g, "hc": d["hc"], "name": stop_name(d)} for g, d in buses[v].items()]
    s.sort(key=lambda x: (-x["hc"], x["name"]))
    bus_stops[v] = s


def analyze(thresh):
    """Greedily group each bus's stops within `thresh` metres into the most-populated anchor."""
    groups, detail, per_bus, items = [], [], [], []
    for v, stops in bus_stops.items():
        assigned, gid, merged_away = set(), 0, 0
        for i, anchor in enumerate(stops):
            if i in assigned:
                continue
            near = [(j, o, hav(anchor["gps"], o["gps"])) for j, o in enumerate(stops)
                    if j != i and j not in assigned and hav(anchor["gps"], o["gps"]) <= thresh]
            if not near:
                continue
            gid += 1
            assigned.add(i)
            for j, _, _ in near:
                assigned.add(j)
            merged_away += len(near)
            items.append({
                "bus": v, "group": gid,
                "target": {"name": anchor["name"], "lat": anchor["gps"][0], "lng": anchor["gps"][1], "riders": anchor["hc"]},
                "merged": [{"name": o["name"], "lat": o["gps"][0], "lng": o["gps"][1], "riders": o["hc"], "dist": round(d)}
                           for j, o, d in sorted(near, key=lambda x: -x[1]["hc"])],
                "combined": anchor["hc"] + sum(o["hc"] for _, o, _ in near),
                "farthest": round(max(d for _, _, d in near)),
            })
            groups.append({
                "bus": v, "unit": unit_of.get(v, ""), "own": type_of.get(v, ""), "group": gid,
                "keep_stop": anchor["name"], "keep_lat": anchor["gps"][0], "keep_lng": anchor["gps"][1],
                "keep_riders": anchor["hc"], "stops_merged_in": len(near),
                "merged_stops": "; ".join(o["name"] for _, o, _ in near),
                "farthest_m": round(max(d for _, _, d in near)),
                "combined_riders": anchor["hc"] + sum(o["hc"] for _, o, _ in near),
            })
            detail.append({"bus": v, "group": gid, "action": "KEEP (target)", "stop": anchor["name"],
                           "lat": anchor["gps"][0], "lng": anchor["gps"][1], "riders": anchor["hc"],
                           "dist_to_target_m": 0, "target_stop": anchor["name"]})
            for j, o, dist in sorted(near, key=lambda x: -x[1]["hc"]):
                detail.append({"bus": v, "group": gid, "action": "→ merge into target", "stop": o["name"],
                               "lat": o["gps"][0], "lng": o["gps"][1], "riders": o["hc"],
                               "dist_to_target_m": round(dist), "target_stop": anchor["name"]})
        per_bus.append({"bus": v, "unit": unit_of.get(v, ""), "stops_before": len(stops),
                        "merge_groups": gid, "stops_merged_away": merged_away,
                        "stops_after": len(stops) - merged_away})
    return groups, detail, per_bus, items


results = {t: analyze(t) for t in THRESHOLDS}

# ---------------- JSON for the Prev-route "Merge review" tab (always written) ----------------
import os
os.makedirs("public", exist_ok=True)
tot_before = sum(b["stops_before"] for b in results[THRESHOLDS[0]][2])
merge_json = {
    "meta": {"date": latest, "stops_asis": tot_before, "buses": len(bus_stops),
             "depot": {"lat": DEPOT[0], "lng": DEPOT[1]}, "bands": {"safe": 150, "watch": 250}},
    "thresholds": {},
}
for _t in THRESHOLDS:
    _g, _d, _pb, _items = results[_t]
    _away = sum(b["stops_merged_away"] for b in _pb)
    merge_json["thresholds"][str(int(_t))] = {
        "groups": len(_items), "stops_merged_away": _away, "remaining": tot_before - _away,
        "riders_moved": sum(sum(m["riders"] for m in it["merged"]) for it in _items),
        "items": _items,
    }
json.dump(merge_json, open(OUT_JSON, "w"))
print(f"Wrote {OUT_JSON} · " + " · ".join(
    f"{k}m {v['groups']}grp/{v['stops_merged_away']}merged" for k, v in merge_json["thresholds"].items()))
if ARGS.json_only:
    raise SystemExit

# ---------------- xlsx (supervisor workbook) ----------------
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------- styling helpers ----------------
HEAD = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="4F46E5")
HFILL2 = PatternFill("solid", fgColor="0EA5E9")
KEEP_FILL = PatternFill("solid", fgColor="E7F5EF")
THIN = Side(style="thin", color="D9DEE6")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center")
TITLE = Font(bold=True, size=14)


def table(ws, headers, keys, records, widths, num_fmts=None, r0=1, fill=HFILL):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r0, c, h)
        cell.font, cell.fill, cell.alignment, cell.border = HEAD, fill, CEN, BORD
    for rix, rec in enumerate(records, r0 + 1):
        for c, key in enumerate(keys, 1):
            cell = ws.cell(rix, c, rec.get(key, ""))
            cell.border = BORD
            if num_fmts and key in num_fmts:
                cell.number_format = num_fmts[key]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return r0 + 1 + len(records)


wb = openpyxl.Workbook()

# ---------------- Sheet 1: Summary / comparison ----------------
ws = wb.active
ws.title = "Summary"
ws["A1"] = "Stop merge suggestions — 200 m vs 300 m"
ws["A1"].font = TITLE
tot_before = sum(b["stops_before"] for b in results[THRESHOLDS[0]][2])
meta = [
    ("ERP date", latest),
    ("Rule", "Merge nearby stops INTO the more populated stop; per bus/route"),
    ("Buses analysed", len(bus_stops)),
    ("Total stops (as-is, un-merged)", tot_before),
    ("NOTE", "Suggestions only — no data changed. For supervisor approval."),
]
for i, (k, val) in enumerate(meta, 3):
    ws.cell(i, 1, k).font = Font(bold=True)
    ws.cell(i, 2, val)
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 30

# side-by-side totals
cmp_r0 = 3 + len(meta) + 1
ws.cell(cmp_r0 - 1, 1, "Threshold comparison").font = Font(bold=True, size=12)
comp_rows = []
for t in THRESHOLDS:
    _, _, pb, _ = results[t]
    away = sum(b["stops_merged_away"] for b in pb)
    comp_rows.append({
        "metric": f"{int(t)} m", "groups": sum(b["merge_groups"] for b in pb),
        "away": away, "after": tot_before - away,
        "reduction": round(away / tot_before * 100, 1) if tot_before else 0,
    })
table(ws, ["Threshold", "Merge groups", "Stops merged away", "Stops after merge", "% reduction"],
      ["metric", "groups", "away", "after", "reduction"], comp_rows,
      [14, 16, 18, 18, 14], {"reduction": "0.0"}, r0=cmp_r0)

# per-bus comparison
pb_r0 = cmp_r0 + len(comp_rows) + 3
ws.cell(pb_r0 - 1, 1, "Per-bus comparison").font = Font(bold=True, size=12)
by_bus = {b["bus"]: b for b in results[THRESHOLDS[0]][2]}
pb200 = {b["bus"]: b for b in results[200.0][2]}
pb300 = {b["bus"]: b for b in results[300.0][2]}
pb_rows = []
for v in sorted(by_bus):
    pb_rows.append({
        "bus": v, "unit": by_bus[v]["unit"], "before": by_bus[v]["stops_before"],
        "g2": pb200[v]["merge_groups"], "a2": pb200[v]["stops_merged_away"], "af2": pb200[v]["stops_after"],
        "g3": pb300[v]["merge_groups"], "a3": pb300[v]["stops_merged_away"], "af3": pb300[v]["stops_after"],
        "extra": pb300[v]["stops_merged_away"] - pb200[v]["stops_merged_away"],
    })
pb_rows.sort(key=lambda x: -x["extra"])
table(ws, ["Bus", "Unit", "Stops as-is", "200m groups", "200m merged", "200m after",
           "300m groups", "300m merged", "300m after", "Extra merged (300 vs 200)"],
      ["bus", "unit", "before", "g2", "a2", "af2", "g3", "a3", "af3", "extra"], pb_rows,
      [12, 11, 11, 12, 12, 11, 12, 12, 11, 22], r0=pb_r0)
ws.freeze_panes = ws.cell(pb_r0 + 1, 1)

# ---------------- per-threshold sheets ----------------
SUG_H = ["Bus", "Unit", "Own/Rent", "Group #", "KEEP stop (target)", "Target lat", "Target lng",
         "Target riders", "# stops merged in", "Stops merged in", "Farthest stop (m)", "Combined riders"]
SUG_K = ["bus", "unit", "own", "group", "keep_stop", "keep_lat", "keep_lng",
         "keep_riders", "stops_merged_in", "merged_stops", "farthest_m", "combined_riders"]
DET_H = ["Bus", "Group #", "Action", "Stop", "Lat", "Lng", "Riders", "Dist to target (m)", "Target stop"]
DET_K = ["bus", "group", "action", "stop", "lat", "lng", "riders", "dist_to_target_m", "target_stop"]

for t in THRESHOLDS:
    groups, detail, _, _ = results[t]
    fill = HFILL if t == 200.0 else HFILL2
    wsg = wb.create_sheet(f"Suggestions {int(t)}m")
    table(wsg, SUG_H, SUG_K, sorted(groups, key=lambda g: (g["bus"], g["group"])),
          [12, 11, 9, 8, 24, 11, 11, 12, 15, 40, 15, 14],
          {"keep_lat": "0.000000", "keep_lng": "0.000000"}, fill=fill)
    wsg.freeze_panes = "A2"
    wsg.auto_filter.ref = f"A1:{get_column_letter(len(SUG_H))}{len(groups)+1}"

    wsd = wb.create_sheet(f"Detail {int(t)}m")
    table(wsd, DET_H, DET_K, sorted(detail, key=lambda d: (d["bus"], d["group"], d["dist_to_target_m"])),
          [12, 8, 18, 22, 11, 11, 9, 16, 22], {"lat": "0.000000", "lng": "0.000000"}, fill=fill)
    wsd.freeze_panes = "A2"
    for rix in range(2, wsd.max_row + 1):
        if "KEEP" in str(wsd.cell(rix, 3).value or ""):
            for c in range(1, len(DET_H) + 1):
                wsd.cell(rix, c).fill = KEEP_FILL

wb.save(OUT)
print(f"ERP date {latest} · {len(bus_stops)} buses · {tot_before} stops as-is")
for t in THRESHOLDS:
    _, _, pb, _ = results[t]
    away = sum(b["stops_merged_away"] for b in pb)
    print(f"  {int(t)} m: {sum(b['merge_groups'] for b in pb)} groups · {away} merged away · "
          f"{tot_before - away} remain ({round(away/tot_before*100,1)}% fewer)")
print(f"Wrote {OUT}")
