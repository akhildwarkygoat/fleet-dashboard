#!/usr/bin/env python3
"""Every stop, per service, with the buses that serve it — as an Excel workbook.

Built to identify stops on the ground: coordinates, a Google Maps link, the riders who
board there, their departments, and which bus already picks them up. Also flags whether
the stop exists in the routed network (data/bus_stops.csv) or would be a NEW node in a
distance-matrix build.

Stops are UNMERGED: one row per distinct home coordinate. Riders sharing an identical
coordinate (same house) share a stop.

Usage:  python build_stops_sheet.py <stops_export.json> <out.xlsx>
"""
import json, sys, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "stops_export.json"
OUT = sys.argv[2] if len(sys.argv) > 2 else "Stops by service.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
NOTE = Font(name=FONT, italic=True, size=9, color="7F7F7F")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
NEW_FILL = PatternFill("solid", fgColor="FFF3CD")     # amber — not in the routed network
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SERVICE_FILL = {
    "9 am General": "DCE9F7", "7 am Morning": "FBE6CC",
    "Rotational": "D6EFEA", "Zenwear": "F7D9E3",
}


def header(ws, cols, row=1):
    for i, (title, width) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=title)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def main():
    d = json.load(open(SRC, encoding="utf-8"))
    stops, buses, day = d["stops"], d["buses"], d["day"]
    no_gps = d.get("noGps", [])
    services = []
    for s in stops:
        if s["service"] not in services:
            services.append(s["service"])

    wb = Workbook()

    # ---------------------------------------------------------------- README
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 104
    ws["A1"] = "Stops by service — identification sheet"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    new_n = sum(1 for s in stops if s["isNew"])
    meta = [
        ("Feed date", day + "  (newest day in the ERP punch feed)"),
        ("Source", "POST http://life.gainup.in:8089/api/general/VehicleEmpMapDetails — rider home GPS"),
        ("Stops", f"{len(stops):,} across {len(services)} services"),
        ("Riders placed", f"{sum(s['headcount'] for s in stops):,}"),
        ("Not routed yet", f"{new_n:,} stops have no node within 200 m in data/bus_stops.csv — amber rows"),
        ("", ""),
        ("HOW STOPS WERE MADE", ""),
        ("Unmerged", "One row per DISTINCT home coordinate. Nothing is collapsed by radius — only riders "
                     "at an identical coordinate (same house) share a stop. Merging is a walking-distance "
                     "decision and has deliberately not been applied."),
        ("Routed vs new", "\"In routed network\" = the curated 739-stop set the optimiser already plans on "
                          "has a stop within 200 m. \"NEW\" = it does not, so a distance-matrix build must "
                          "add this node. NEW does not mean the stop is missing — it exists and has riders."),
        ("Depot", "Every service runs from the Batlagundu factory except Zenwear, which runs from its own "
                  "Subbulapuram site (9.6732711, 77.8072837). Distances are measured from the right one."),
        ("", ""),
        ("HOW TO IDENTIFY A STOP", ""),
        ("Map link", "Every row has a Google Maps link on its coordinates — the fastest way to see the place."),
        ("Bus", "The bus(es) that already carry riders from that stop, with how many each. This is the ERP's "
                "standing assignment, not a per-trip route."),
        ("Riders", "The people who board there, and their departments. Use these to ask someone on the ground."),
        ("Village", "The locality the ERP reports for those riders. Some read as ????? — the ERP serves "
                    "non-ASCII place names in the wrong character set, so Tamil names are unreadable at source."),
        ("", ""),
        ("CAVEAT", ""),
        ("Riders with no GPS", f"{len(no_gps)} employees have no home coordinates in the ERP, so they have NO stop "
                               "and appear on no route. They are listed on the 'Riders without GPS' sheet."),
    ]
    r = 3
    for k, v in meta:
        if k and not v:
            ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=11, color="1F3864")
        elif k:
            ws.cell(row=r, column=1, value=k).font = BOLD
            c = ws.cell(row=r, column=2, value=v)
            c.font, c.alignment = BODY, Alignment(wrap_text=True, vertical="top")
        r += 1

    # -------------------------------------------------------------- Summary
    ws = wb.create_sheet("Summary")
    header(ws, [("Service", 18), ("Depot", 26), ("Stops", 9), ("Riders", 9),
                ("In routed network", 17), ("NEW stops", 11), ("Riders at new stops", 18), ("Buses", 8)])
    for i, sv in enumerate(services, start=2):
        rows_ = [s for s in stops if s["service"] == sv]
        nu = [s for s in rows_ if s["isNew"]]
        vehs = {b for s in rows_ for b, _ in s["buses"]}
        vals = [sv, rows_[0]["depot"], len(rows_), sum(s["headcount"] for s in rows_),
                len(rows_) - len(nu), len(nu), sum(s["headcount"] for s in nu), len(vehs)]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font, c.border = BODY, BOX
            if col == 1:
                c.font = BOLD
                c.fill = PatternFill("solid", fgColor=SERVICE_FILL.get(sv, "F2F2F2"))
    last = len(services) + 1
    ws.cell(row=last + 2, column=1, value="TOTAL").font = BOLD
    for col in (3, 4, 5, 6, 7):
        c = ws.cell(row=last + 2, column=col,
                    value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{last})")
        c.font = BOLD
    ws.cell(row=last + 4, column=1,
            value="A NEW stop is not a missing stop — it exists and has riders, it just has no node in the "
                  "distance matrix yet. Amber rows on the service sheets are the ones a matrix build must add.").font = NOTE

    # ------------------------------------------------- one sheet per service
    for sv in services:
        rows_ = sorted([s for s in stops if s["service"] == sv], key=lambda x: -x["headcount"])
        ws = wb.create_sheet(sv[:31])
        header(ws, [("#", 5), ("Stop / village", 30), ("Riders", 8), ("Bus(es)", 22),
                    ("Latitude", 12), ("Longitude", 12), ("Map", 9), ("From depot", 11),
                    ("In routed network?", 17), ("Nearest routed stop", 17),
                    ("Departments", 26), ("Who boards here", 60)])
        for i, s in enumerate(rows_, start=2):
            bus_txt = ", ".join(f"{b} ({n})" for b, n in s["buses"]) or "—"
            vals = [i - 1, s["name"] or s["village"] or "—", s["headcount"], bus_txt,
                    round(s["lat"], 6), round(s["lng"], 6), None,
                    s["depotKm"], "NEW — not routed" if s["isNew"] else "yes",
                    "—" if s["isNew"] else f'{s["nearestExistingM"]} m',
                    ", ".join(d for d in s["depts"] if d) or "—",
                    ", ".join(s["riderNames"][:14]) + (" …" if len(s["riderNames"]) > 14 else "")]
            for col, v in enumerate(vals, start=1):
                c = ws.cell(row=i, column=col, value=v)
                c.font, c.border = BODY, BOX
                if s["isNew"]:
                    c.fill = NEW_FILL
                if col in (5, 6):
                    c.number_format = "0.000000"
                if col in (1, 3, 8):
                    c.number_format = "0.#"
            # clickable pin — the quickest way to place a stop on the ground
            link = ws.cell(row=i, column=7, value="Open")
            link.hyperlink = f'https://www.google.com/maps/search/?api=1&query={s["lat"]:.6f},{s["lng"]:.6f}'
            link.font = Font(name=FONT, size=10, color="0563C1", underline="single")
            link.border = BOX
            if s["isNew"]:
                link.fill = NEW_FILL
            ws.cell(row=i, column=9).font = Font(name=FONT, size=10, bold=s["isNew"],
                                                 color="9C6500" if s["isNew"] else "1F7A46")
        ws.auto_filter.ref = f"A1:L{len(rows_) + 1}"
        n = len(rows_) + 3
        ws.cell(row=n, column=2, value=f"{len(rows_)} stops · {sum(s['headcount'] for s in rows_)} riders · "
                                      f"{sum(1 for s in rows_ if s['isNew'])} not in the routed network "
                                      f"(amber). Depot: {rows_[0]['depot']}.").font = NOTE

    # ------------------------------------------------------- Bus → its stops
    ws = wb.create_sheet("By bus")
    header(ws, [("Bus", 14), ("Unit", 13), ("Type", 16), ("Seats", 8), ("Service(s)", 30),
                ("Stops", 8), ("Riders", 8), ("Stops served (most riders first)", 90)])
    by_bus = collections.defaultdict(list)
    for s in stops:
        for b, n in s["buses"]:
            by_bus[b].append((s, n))
    meta_bus = {b["vehicle"]: b for b in buses}
    for i, b in enumerate(sorted(by_bus), start=2):
        items = sorted(by_bus[b], key=lambda x: -x[1])
        m = meta_bus.get(b, {})
        svcs = sorted({s["service"] for s, _ in items})
        vals = [b, m.get("unit", "—"), m.get("type", "—"), m.get("capacity", ""),
                ", ".join(svcs), len(items), sum(n for _, n in items),
                "; ".join(f'{(s["name"] or s["village"] or "?")} ({n})' for s, n in items[:22])
                + (" …" if len(items) > 22 else "")]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font, c.border = BODY, BOX
            if col == 1:
                c.font = BOLD
    ws.auto_filter.ref = f"A1:H{len(by_bus) + 1}"
    ws.cell(row=len(by_bus) + 3, column=2,
            value="A bus can appear under more than one service — the ERP assigns riders to a bus, not to a "
                  "trip, so one vehicle covers several runs across the day.").font = NOTE

    # ------------------------------------------------- riders with no GPS
    ws = wb.create_sheet("Riders without GPS")
    header(ws, [("Employee no.", 14), ("Code", 12), ("Name", 30), ("Department", 20),
                ("Unit", 13), ("Shift", 22), ("Bus", 14), ("Stops on their bus", 9),
                ("Likely stop — which of these is theirs?", 78)])
    # Their bus is known, and its other riders DO have GPS. So the question is not "where do
    # they live?" but "which of this bus's existing stops is theirs?" — a much shorter one.
    stops_by_bus = collections.defaultdict(list)
    for s in stops:
        for b, n in s["buses"]:
            stops_by_bus[b].append((s, n))
    no_gps_sorted = sorted(no_gps, key=lambda r: (r["bus"] or "", r["name"] or ""))
    for i, r_ in enumerate(no_gps_sorted, start=2):
        cands = sorted(stops_by_bus.get(r_["bus"], []), key=lambda x: -x[1])
        names = [f'{(c["name"] or c["village"] or "?")} ({n})' for c, n in cands[:12]]
        vals = [r_["empl"], r_["code"], r_["name"], r_["dept"], r_["unit"], r_["shift"], r_["bus"],
                len(cands), "; ".join(names) + (" …" if len(cands) > 12 else "")]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font, c.border = BODY, BOX
            if col == 7:
                c.font = BOLD
    ws.auto_filter.ref = f"A1:I{len(no_gps) + 1}"
    ws.cell(row=len(no_gps) + 3, column=2,
            value="These employees have no home coordinates in the ERP, so they have no stop and appear on no "
                  "route. Their bus IS known, and every other rider on it has GPS — so the last column lists that "
                  "bus's existing stops (rider counts in brackets), most-used first. Confirming which one each "
                  "person boards at is usually enough; only a rider who boards somewhere new needs a fresh "
                  "coordinate. All 69 sit on 25 buses, so this is 25 conversations, not 69.").font = NOTE

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  {len(stops):,} stops · {sum(s['headcount'] for s in stops):,} riders · "
          f"{new_n:,} not in the routed network · {len(no_gps)} riders with no GPS")


if __name__ == "__main__":
    main()
