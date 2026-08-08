#!/usr/bin/env python3
"""Riders with no home coordinates in the ERP — the collection list.

Every one of them is already allocated to a bus, and that bus's other riders DO have
GPS. So the question is rarely "where do they live?" but "which of this bus's existing
stops do they board at?" — the sheet leads with that.

Usage:  python build_missing_gps_sheet.py <missing_gps.json> <out.xlsx>
"""
import json, sys, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "missing_gps.json"
OUT = sys.argv[2] if len(sys.argv) > 2 else "Riders missing coordinates.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
NOTE = Font(name=FONT, italic=True, size=9, color="7F7F7F")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
NEW_FILL = PatternFill("solid", fgColor="FFF3CD")     # amber — appeared since the last check
FILL_IN = PatternFill("solid", fgColor="E8F1DE")      # green — columns for you to complete
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header(ws, cols, row=1):
    for i, (title, width) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=title)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def main():
    d = json.load(open(SRC, encoding="utf-8"))
    rows, day = d["missing"], d["day"]
    by_bus = collections.defaultdict(list)
    for r in rows:
        by_bus[r["bus"]].append(r)

    wb = Workbook()

    # ---------------------------------------------------------------- README
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 104
    ws["A1"] = "Riders with no home coordinates"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    outstanding = sum(1 for r in rows if r["since"] == "outstanding")
    for i, (k, v) in enumerate([
        ("Feed date", f"{day}  (newest day in the ERP punch feed)"),
        ("Riders affected", f"{len(rows)} — {outstanding} outstanding from the last check, "
                            f"{len(rows) - outstanding} new since 06-08-2026"),
        ("Buses involved", f"{len(by_bus)}"),
        ("", ""),
        ("WHY IT MATTERS", ""),
        ("Not routable", "All of them are allocated to a bus and most travel daily — but with no home "
                         "coordinate the optimiser cannot place them, so they appear on no route and are "
                         "silently excluded from every plan."),
        ("Costs money later", "A coordinate added before the distance-matrix build rides along for free. "
                              "One added afterwards may need its own matrix node, bought separately."),
        ("", ""),
        ("HOW TO FILL IT IN", ""),
        ("Start with the bus", "Sort by Bus. Each bus's other riders already have GPS, so the 'Likely stop' "
                               "column lists the stops that bus already serves, busiest first. Usually the "
                               "answer is one of those — no new address needed."),
        ("Only if none fit", "If the rider boards somewhere the bus does not already stop, capture a real "
                             "coordinate. The two green columns are there for that."),
        ("Where it goes", "Enter the result in the ERP, not this sheet. The dashboard re-derives everything "
                          "from the feed on each sync, so an ERP update appears automatically."),
        ("", ""),
        ("COLUMN KEY", ""),
        ("Amber row", "Appeared since the 06-08 check — a newer gap."),
        ("Green columns", "Blank, for you to complete."),
        ("Travelled today", "Whether they actually rode on the feed date — a 'no' may be leave, not a bad record."),
    ], start=3):
        if k and not v:
            ws.cell(row=i, column=1, value=k).font = Font(name=FONT, bold=True, size=11, color="1F3864")
        elif k:
            ws.cell(row=i, column=1, value=k).font = BOLD
            c = ws.cell(row=i, column=2, value=v)
            c.font, c.alignment = BODY, Alignment(wrap_text=True, vertical="top")

    # ------------------------------------------------------------ the riders
    ws = wb.create_sheet("Riders to collect")
    header(ws, [("Employee no.", 13), ("Code", 11), ("Name", 28), ("Department", 18),
                ("Service", 15), ("Unit", 12), ("Bus", 13), ("Travelled today", 13),
                ("Status", 15), ("Stops on their bus", 9),
                ("Likely stop — which of these is theirs?", 66),
                ("Stop they board at", 26), ("New coordinate (lat, lng)", 24)])
    ordered = sorted(rows, key=lambda r: (r["bus"] or "", r["name"] or ""))
    for i, r in enumerate(ordered, start=2):
        cands = "; ".join(f'{c["name"]} ({c["n"]})' for c in r["candidates"]) or "—"
        vals = [r["empl"], r["code"], r["name"], r["dept"], r["service"], r["unit"], r["bus"],
                "yes" if r["present"] else "no", r["since"], r["busStopCount"], cands, None, None]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font, c.border = BODY, BOX
            if col == 7:
                c.font = BOLD
            if r["since"] != "outstanding" and col <= 11:
                c.fill = NEW_FILL
            if col in (12, 13):
                c.fill = FILL_IN
    ws.auto_filter.ref = f"A1:M{len(ordered) + 1}"
    ws.cell(row=len(ordered) + 3, column=3,
            value="Green columns are for you. Usually the answer is already in 'Likely stop' — the bus "
                  "serves that village and the rider boards at one of its existing stops; only write a new "
                  "coordinate when none of them fit.").font = NOTE

    # -------------------------------------------------------- bus-by-bus view
    ws = wb.create_sheet("By bus")
    header(ws, [("Bus", 14), ("Riders to collect", 15), ("Stops that bus serves", 9),
                ("Who", 46), ("Stops it already serves", 74)])
    for i, b in enumerate(sorted(by_bus, key=lambda x: -len(by_bus[x])), start=2):
        grp = by_bus[b]
        cands = grp[0]["candidates"]
        vals = [b, len(grp), grp[0]["busStopCount"],
                ", ".join((g["name"] or "")[:18] for g in grp),
                "; ".join(f'{c["name"]} ({c["n"]})' for c in cands) or "— none known"]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font, c.border = BODY, BOX
            if col == 1:
                c.font = BOLD
    ws.auto_filter.ref = f"A1:E{len(by_bus) + 1}"
    ws.cell(row=len(by_bus) + 3, column=2,
            value="Work down this sheet rather than the rider list — one conversation per bus clears "
                  "several riders at once.").font = NOTE

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  {len(rows)} riders · {len(by_bus)} buses · {outstanding} outstanding, {len(rows)-outstanding} new")


if __name__ == "__main__":
    main()
