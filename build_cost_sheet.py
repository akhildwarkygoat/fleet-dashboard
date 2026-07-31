#!/usr/bin/env python3
"""Recreate the ERP's vehicle costing data as an Excel workbook.

The ERP exposes costs on the SAME endpoint as the employee/attendance feed
(POST /api/general/VehicleEmpMapDetails), joined so that every cost line is
repeated once per employee riding that vehicle. This script strips that
duplication and lays the data out as it is actually modelled:

  Description   what the cost is for      (ROAD TAX, TYRE, FC WORK, ...)
  Period_Name   the period it applies to  (YEARLY, QUARTER - I, MARCH, ...)
  Bud_AMt       the amount for THAT line  (varies per line)
  VMS_Cost      a per-vehicle total       (one value per vehicle, repeated)

Usage:  python build_cost_sheet.py <erp.json> <out.xlsx>
"""
import json, re, sys, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "/tmp/erp_now.json"
OUT = sys.argv[2] if len(sys.argv) > 2 else "ERP_vehicle_costs.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
NOTE_FONT = Font(name=FONT, italic=True, size=9, color="7F7F7F")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
MONEY = '#,##0.00;(#,##0.00);-'
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(path):
    """Pull cost lines out of the feed. Tolerates a truncated download: rows are
    matched individually, so a half-written final row is simply skipped."""
    raw = open(path, encoding="utf-8", errors="ignore").read()
    rows = re.findall(r"\{[^{}]*\}", raw)
    get = lambda r, k: (re.search(r'"%s":"([^"]*)"' % k, r) or [None, None])[1]

    lines, veh, employees, dates = {}, {}, collections.defaultdict(set), set()
    for r in rows:
        v = get(r, "VehName")
        if not v:
            continue
        d = get(r, "date")
        if d:
            dates.add(d[:10])
        employees[v].add(get(r, "Empl_no"))
        info = veh.setdefault(v, {"Type": get(r, "Type"), "Seat": get(r, "Seat"),
                                  "Mileage": get(r, "Mileage"), "Compname": get(r, "Compname"),
                                  "VMS_Cost": None, "Month": None, "Year": None})
        if get(r, "Year"):                      # only cost rows carry Month/Year
            info["Month"], info["Year"] = get(r, "Month"), get(r, "Year")
        c = get(r, "VMS_Cost")
        if c and float(c or 0) != 0:
            info["VMS_Cost"] = float(c)
        desc, amt, per = get(r, "Description"), get(r, "Bud_AMt"), get(r, "Period_Name")
        if desc and desc != "-":
            # dedupe the per-employee repetition; identical lines collapse to one
            lines[(v, desc, per, amt)] = {"VehName": v, "Description": desc,
                                          "Period_Name": per, "Bud_AMt": float(amt or 0)}
    return rows, lines, veh, employees, dates


def header(ws, cols, row=1):
    for i, (title, width) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=title)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def main():
    rows, lines, veh, employees, dates = load(SRC)
    recs = sorted(lines.values(), key=lambda r: (r["VehName"], r["Description"], r["Period_Name"] or ""))
    # the feed's own year, so "years since stamp" stays right when this is re-run later
    FEED_YEAR = max((int(d[-4:]) for d in dates if d[-4:].isdigit()), default=2026)
    vehicles = sorted(veh)
    descs = sorted({r["Description"] for r in recs})
    periods = sorted({r["Period_Name"] for r in recs if r["Period_Name"]})

    wb = Workbook()

    # ---------------------------------------------------------------- README
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 104
    ws["A1"] = "ERP vehicle costing data"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    meta = [
        ("Source", "POST http://172.16.10.169:8089/api/general/VehicleEmpMapDetails  (body {})"),
        ("Feed date(s)", ", ".join(sorted(dates)[:6]) + (" …" if len(dates) > 6 else "")),
        ("Rows in feed", f"{len(rows):,} raw rows"),
        ("Unique cost lines", f"{len(recs):,} after removing the per-employee duplication"),
        ("Redundancy", f"{len(rows)/max(1,len(recs)):.0f}x — every cost line is repeated once per employee on that vehicle"),
        ("Vehicles", f"{len(vehicles)} in this extract"),
        ("", ""),
        ("HOW THE ERP MODELS IT", ""),
        ("Description", "What the cost is for. " + str(len(descs)) + " distinct values, e.g. " + ", ".join(descs[:5])),
        ("Period_Name", "Period the line applies to: " + ", ".join(periods[:8])),
        ("Bud_AMt", "The amount for THAT cost line. Varies line to line — this is the real per-line cost."),
        ("VMS_Cost", "A per-vehicle figure, one value repeated on every row for that vehicle. NOT a per-line amount, so it must never be summed across lines."),
        ("Month / Year", "Present on cost rows only, but ONE stamp per vehicle — every cost line on a bus carries the same value, and no vehicle has more than one. It therefore does NOT date the individual costs. Values run 2010-2026 while the feed's attendance date is 2026, so it reads as the vehicle's induction/registration date."),
        ("", ""),
        ("UNRESOLVED", ""),
        ("Reconciliation", "Summing a vehicle's Bud_AMt lines does not equal its VMS_Cost for any vehicle. See the Reconciliation sheet. Which figure is authoritative, and what period VMS_Cost covers, is an open question for the ERP team."),
        ("Coverage", "Built from a download that the server had not finished streaming, so a vehicle's later cost lines may be missing and line sums may be understated. Re-run this script against a complete response to refresh."),
        ("Field names", "Bud_AMt carries costs despite reading like a budget; VMS_Cost carries a total. Names are the ERP's, reproduced unchanged."),
        ("No date per cost line", "Nothing in the feed says which year a given payment belongs to. That is why the duplicate pairs cannot be judged from the data alone: two ROAD TAX QUARTER - I rows Rs 10 apart could be two years or one typo, and the feed cannot tell you which."),
    ]
    r = 3
    for k, v in meta:
        if k and not v:
            ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=11, color="1F3864")
        else:
            ws.cell(row=r, column=1, value=k).font = BOLD
            c = ws.cell(row=r, column=2, value=v)
            c.font = BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # ------------------------------------------------------------- CostLines
    ws = wb.create_sheet("CostLines")
    header(ws, [("VehName", 14), ("Type", 17), ("Seat", 6), ("Compname", 26),
                ("Description", 28), ("Period_Name", 15), ("Bud_AMt", 14), ("VMS_Cost (vehicle)", 18),
                ("Month (vehicle)", 15), ("Year (vehicle)", 14)])
    for i, rec in enumerate(recs, start=2):
        v = veh[rec["VehName"]]
        for col, val in enumerate([rec["VehName"], v["Type"], v["Seat"], v["Compname"],
                                   rec["Description"], rec["Period_Name"], rec["Bud_AMt"],
                                   v["VMS_Cost"], v["Month"],
                                   int(v["Year"]) if (v["Year"] or "").isdigit() else v["Year"]], start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font, c.border = BODY, BOX
            if col in (7, 8):
                c.number_format = MONEY
    last = len(recs) + 1
    tr = last + 2
    ws.cell(row=tr, column=6, value="Total Bud_AMt").font = BOLD
    t = ws.cell(row=tr, column=7, value=f"=SUM(G2:G{last})")
    t.font, t.number_format = BOLD, MONEY
    ws.cell(row=tr + 1, column=6, value="Lines").font = BOLD
    ws.cell(row=tr + 1, column=7, value=f"=COUNTA(E2:E{last})").font = BOLD
    n = ws.cell(row=tr + 3, column=1,
                value="Each row is one distinct (vehicle, description, period, amount) from the ERP. "
                      "VMS_Cost, Month and Year all repeat per vehicle by design — do not sum columns H-J. "
                      "Month/Year is ONE stamp per vehicle, not the date of the cost on that row: every line "
                      "on a bus carries the same value, so it cannot say which year a payment belongs to.")
    n.font = NOTE_FONT

    # ---------------------------------------------------------- Reconciliation
    ws = wb.create_sheet("Reconciliation")
    header(ws, [("VehName", 14), ("Type", 17), ("Seat", 6), ("Cost lines", 11),
                ("Sum of Bud_AMt", 17), ("VMS_Cost", 16), ("Gap (sum - VMS)", 17), ("Gap %", 10),
                ("Month", 10), ("Year", 8), ("Yrs since stamp", 15)])
    for i, v in enumerate(vehicles, start=2):
        info = veh[v]
        ws.cell(row=i, column=1, value=v).font = BODY
        ws.cell(row=i, column=2, value=info["Type"]).font = BODY
        ws.cell(row=i, column=3, value=info["Seat"]).font = BODY
        ws.cell(row=i, column=4, value=f"=COUNTIFS(CostLines!$A$2:$A${last},$A{i})").font = BODY
        for col, f in ((5, f"=SUMIFS(CostLines!$G$2:$G${last},CostLines!$A$2:$A${last},$A{i})"),
                       (6, info["VMS_Cost"]),
                       (7, f"=IF(F{i}=\"\",\"\",E{i}-F{i})")):
            c = ws.cell(row=i, column=col, value=f)
            c.font, c.number_format, c.border = BODY, MONEY, BOX
        p = ws.cell(row=i, column=8, value=f"=IFERROR(G{i}/F{i},\"\")")
        p.font, p.number_format = BODY, "0.0%"
        ws.cell(row=i, column=9, value=info["Month"]).font = BODY
        yr = int(info["Year"]) if (info["Year"] or "").isdigit() else None
        ws.cell(row=i, column=10, value=yr).font = BODY
        a = ws.cell(row=i, column=11, value=(f"=IF(J{i}=\"\",\"\",{FEED_YEAR}-J{i})" if yr else None))
        a.font, a.number_format = BODY, "0"
    vlast = len(vehicles) + 1
    ws.cell(row=vlast + 2, column=4, value="Totals").font = BOLD
    for col in (5, 6, 7):
        c = ws.cell(row=vlast + 2, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{vlast})")
        c.font, c.number_format = BOLD, MONEY
    ws.cell(row=vlast + 4, column=1,
            value="VMS_Cost tracks the vehicle stamp year, not its running cost: buses stamped 2016 or earlier average "
                  "~Rs 12.2 lakh while those stamped 2022+ average ~Rs 44k, and Feb-2026 buses sit at zero. That points to "
                  "VMS_Cost being spend accumulated since the stamp date, which is why it never equals the cost lines "
                  "and why it is the wrong basis for a per-day cost.").font = NOTE_FONT

    # ------------------------------------------------------------ ByCostType
    ws = wb.create_sheet("ByCostType")
    header(ws, [("Description", 30), ("Lines", 9), ("Vehicles", 10), ("Sum of Bud_AMt", 18), ("Share", 10)])
    for i, d in enumerate(descs, start=2):
        ws.cell(row=i, column=1, value=d).font = BODY
        ws.cell(row=i, column=2, value=f"=COUNTIFS(CostLines!$E$2:$E${last},$A{i})").font = BODY
        nveh = len({r["VehName"] for r in recs if r["Description"] == d})
        ws.cell(row=i, column=3, value=nveh).font = BODY
        c = ws.cell(row=i, column=4, value=f"=SUMIFS(CostLines!$G$2:$G${last},CostLines!$E$2:$E${last},$A{i})")
        c.font, c.number_format = BODY, MONEY
        dlast_placeholder = len(descs) + 1
        s = ws.cell(row=i, column=5, value=f"=IFERROR(D{i}/$D${dlast_placeholder + 2},\"\")")
        s.font, s.number_format = BODY, "0.0%"
    dlast = len(descs) + 1
    ws.cell(row=dlast + 2, column=1, value="Total").font = BOLD
    for col, f in ((2, f"=SUM(B2:B{dlast})"), (4, f"=SUM(D2:D{dlast})")):
        c = ws.cell(row=dlast + 2, column=col, value=f)
        c.font, c.number_format = BOLD, (MONEY if col == 4 else "#,##0")

    # -------------------------------------------------------------- ByPeriod
    ws = wb.create_sheet("ByPeriod")
    header(ws, [("Period_Name", 20), ("Lines", 9), ("Sum of Bud_AMt", 18)])
    for i, p in enumerate(periods, start=2):
        ws.cell(row=i, column=1, value=p).font = BODY
        ws.cell(row=i, column=2, value=f"=COUNTIFS(CostLines!$F$2:$F${last},$A{i})").font = BODY
        c = ws.cell(row=i, column=3, value=f"=SUMIFS(CostLines!$G$2:$G${last},CostLines!$F$2:$F${last},$A{i})")
        c.font, c.number_format = BODY, MONEY
    plast = len(periods) + 1
    ws.cell(row=plast + 2, column=1, value="Total").font = BOLD
    c = ws.cell(row=plast + 2, column=3, value=f"=SUM(C2:C{plast})")
    c.font, c.number_format = BOLD, MONEY
    ws.cell(row=plast + 4, column=1,
            value="Period_Name is the ERP's own vocabulary — YEARLY, QUARTER - I..IV and month names. "
                  "It is not the day/month/year the dashboard normalises with, so a mapping is needed before these "
                  "amounts can become a per-working-day cost.").font = NOTE_FONT

    # ---------------------------------------------------------------- ByYear
    ws = wb.create_sheet("ByYear")
    header(ws, [("Stamp year", 12), ("Vehicles", 10), ("Avg VMS_Cost", 16), ("Total VMS_Cost", 18)])
    yr_groups = collections.defaultdict(list)
    for v in vehicles:
        y = veh[v]["Year"]
        if y and y.isdigit() and veh[v]["VMS_Cost"]:
            yr_groups[int(y)].append(veh[v]["VMS_Cost"])
    for i, y in enumerate(sorted(yr_groups), start=2):
        vals = yr_groups[y]
        ws.cell(row=i, column=1, value=y).font = BODY
        ws.cell(row=i, column=2, value=len(vals)).font = BODY
        for col, val in ((3, sum(vals) / len(vals)), (4, sum(vals))):
            c = ws.cell(row=i, column=col, value=val)
            c.font, c.number_format = BODY, MONEY
    ws.cell(row=len(yr_groups) + 3, column=1,
            value="VMS_Cost falls as the stamp year gets more recent — older buses carry far more, and the newest "
                  "carry zero. That is the signature of spend accumulated since induction, not of a running cost, "
                  "so VMS_Cost should not be turned into a per-day figure.").font = NOTE_FONT

    # ------------------------------------------------------------- RawSample
    ws = wb.create_sheet("RawSample")
    ws.column_dimensions["A"].width = 200
    ws["A1"] = "Verbatim rows from the feed — exactly as the ERP returns them"
    ws["A1"].font = Font(name=FONT, bold=True, size=11)
    ws["A2"] = "Shown so the duplication is visible: the same vehicle's cost line reappears for each employee."
    ws["A2"].font = NOTE_FONT
    shown = 0
    for r in rows:
        if '"Description":"' in r and '"Description":"-"' not in r:
            c = ws.cell(row=4 + shown, column=1, value=r)
            c.font = Font(name="Courier New", size=8)
            c.alignment = Alignment(wrap_text=False)
            shown += 1
            if shown == 12:
                break

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  cost lines {len(recs):,} | vehicles {len(vehicles)} | descriptions {len(descs)} | periods {len(periods)}")
    print(f"  raw rows {len(rows):,} -> {len(rows)/max(1,len(recs)):.0f}x redundancy")


if __name__ == "__main__":
    main()
